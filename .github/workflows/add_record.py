from openpyxl import Workbook, load_workbook
from datetime import datetime

# دالة لإضافة سجل جديد
def add_record(filename, date, profit, balance, zain_cash, taif, notes):
    try:
        workbook = load_workbook(filename)
        sheet = workbook.active
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        headers = ["التاريخ", "الربح", "الرصيد", "زين كاش", "الطيف", "الملاحظات"]
        sheet.append(headers)
    
    new_record = [date, profit, balance, zain_cash, taif, notes]
    sheet.append(new_record)
    workbook.save(filename)

# اسم الملف
filename = 'records.xlsx'

# إضافة سجل جديد
current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
add_record(filename, current_time, 100, 500, 200, 50, "ملاحظات جديدة")

print("تم حفظ السجل بنجاح.")
input("اضغط على Enter للخروج...")
